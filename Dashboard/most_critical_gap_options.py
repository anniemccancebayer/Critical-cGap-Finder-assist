# Import required libraries
import pandas as pd
import dash
from dash import dcc, html, Input, Output, dash_table,no_update
from dash.dependencies import Input, Output, State
import dash_core_components as dcc
import dash_html_components as html
import dash_table
import base64
from io import BytesIO
import dash_bootstrap_components as dbc
import urllib.parse
from dash.exceptions import PreventUpdate
import os
import openpyxl 
from openpyxl.styles import PatternFill, Border, Side
import io
from utils import *

# Initialize the Dash app
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP], suppress_callback_exceptions=True)
# Define the layout of the app

# Load the image
image_filename =  os.path.join(os.path.dirname(__file__),'Logo.png')  # Path to your image file
encoded_image = base64.b64encode(open(image_filename, 'rb').read()).decode('ascii')
# Replace 'your_logo.png' with the path to your company logo
LOGO_URL =  os.path.join(os.path.dirname(__file__),"bayer_logo.png")  # Path to your image file
encoded_image2 = base64.b64encode(open(LOGO_URL, 'rb').read()).decode('ascii')
navbar = dbc.Navbar(
    dbc.Container(
        [
            html.A(
                # Use row and col to control vertical alignment of logo / brand
                dbc.Row(
                    [
                        dbc.Col(html.Img(src=f'data:image/png;base64,{encoded_image2}', height="50px")),
                        dbc.Col(dbc.NavbarBrand(" ", className="ml-2")),
                        #dbc.Col(html.Img(src=f'data:image/png;base64,{encoded_image}', height="55px")),
                    ],
                    align="center",
                 #   no_gutters=True,
                ),
                href="/",  # Link to the home page of your app
            ),
            dbc.NavbarToggler(id="navbar-toggler"),
        ],
        fluid=True,
    ),
    color="light",  # Navbar color
    dark=False,  # Text color
    #className="mb-4",  # Margin bottom
)

# Your app's main content
main_content  = dbc.Container(
    fluid=True,
    style={'backgroundColor': '#e2f1f6', 'padding': '20px'},
    children=[
        dbc.Row(
            dbc.Col(
                html.Img(src=f'data:image/png;base64,{encoded_image}',
                         style={'height': '150px', 'marginBottom': '10px'}), 
                className='text-center'
            ),
        ),


        

        dbc.Row(
            dbc.Col(
                dbc.Card(
                    dbc.CardBody([
                        dcc.Markdown('''
                                        ##### Input File
                                        Excel file "**Master GAP Table with revised GAPs**" provided by the regional regulatory managers,
                                        which includes each of the requested GAPs by country, crop, and product.

                                        ##### Crop Selection
                                        For the purpose of analysis, certain crops have been excluded (rye, triticale, spelt, oat), and some crops have been grouped together, e.g., Barley (spring & winter), Wheat (durum, spring, winter), Cabbage, Onion, Rape.

                                        ##### Critical GAP Criteria
                                     The app identifies, among all these GAPs, the most critical GAP based on:
                                      Formulation (J-neck), Regulatory zone (G-collar), Crop (O-collar)
                                     
                                        The determination of the most critical GAP is based on the following criteria:
                                        - 1. **Application rate (g/ha):** The higher, the more critical.
                                        - 2. **BBCH stage:** The latest stage is considered the most critical.
                                        - 3. **Pre-Harvest Interval (PHI):** The shorter, the more critical.
                                        - 4. **Interval between applications:** The smaller the interval, the more critical.
                                        - 5. **Number of applications:** More applications indicate a higher criticality.

                                        ''') ,
                    ]),
                    className="mb-3 shadow",
                    style={'backgroundColor': '#f5fafc', 'borderRadius': '15px', 'padding': '10px'}
                ),
                width={'size':10, 'offset': 1}  # Center the card on the page
            )
        ),


        dbc.Row(
            dbc.Col(
                dcc.Upload(
                    id='upload-data',
                    children=html.Div([
                        'Drag and Drop or ',
                        html.A('Select Files')
                    ]),
                    style={
                        'width': '100%',
                        'height': '60px',
                        'lineHeight': '60px',
                        'borderWidth': '1px',
                        'borderStyle': 'dashed',
                        'borderRadius': '5px',
                        'textAlign': 'center'
                    },
                    multiple=False
                ),
                width=12
            )
        ),
              
            dbc.Row(
                dbc.Col(
                    html.Div(id='msg_table', style={'marginTop': '30px'}),
                    #width={'size': 8, 'offset': 3}
                )
            ),
    

            

            dbc.Row(
            dbc.Col(
                dcc.Loading(  # Add the Loading component here
                    id="loading",
                    type="default",  # You can choose 'default', 'circle', or 'square'                 
                    children=[
                        html.Div(
                            id='dropdown-container',
                            children=[
                                dcc.Dropdown(
                                    id='regulatory-filter',
                                    options=[],
                                    multi=False,
                                    placeholder='Select Master Country Zone Block',
                                    style={'width': '550px','marginBottom': '10px'},
                                    className='dropdown-custom shadow'
                                ),
                                dcc.Dropdown(
                                    id='ApplicationRate-filter',
                                    options=[],
                                    multi=False,
                                    placeholder='Select the AS dose rate expressions',
                                    style={'width': '550px','marginBottom': '10px'},
                                    className='dropdown-custom shadow'
                                ),
                                
                                dcc.Dropdown(
                                    id='product-filter',
                                    options=[],
                                    multi=True,
                                    placeholder='Select Product',
                                    className='dropdown-custom shadow'
                                ),

                                dcc.Dropdown(
                                    id='crop-filter',
                                    options=[],
                                    multi=True,
                                    placeholder='Select crop'
                                ),
                                
                                dcc.Dropdown(
                                    id='region-filter',
                                    options=[],
                                    multi=True,
                                    placeholder='Select region',
                                    className='dropdown-custom shadow'
                                ),
                                html.Div(
                                    id='filtered-table',
                                    style={'marginTop': '30px'},
                                    className='dropdown-custom shadow'
                                ),



                            ]
                        )
                    ]
                ),
                # width={'size': 8, 'offset': 3}
            )
        ),
        
            # Add the download button and link container
            html.Div([
                html.Div(id='download-link-container'),
                dbc.Button('Export Data', id='download-button', color='primary', className='mt-3'),
                dcc.Download(id='download-dataframe-csv')
                
            ], className='text-center'),

           html.Hr(),  # This creates a horizontal line 
           
        dbc.Row(
            dbc.Col(
                dbc.Card(
                    dbc.CardBody([
                        dcc.Markdown('''
                                         Compare Versions of the Master GAP Table:
                                     
                                      Input another Excel file of the "**Master GAP Table with revised GAPs**"
                                         and analyse the Changes Between GAP Table Versions

                                        ''',style={'textAlign': 'center'})  # Center the text
                    ]),
                    className="mb-3 shadow",
                    style={'backgroundColor': '#e2f1f6', 'borderRadius': '15px', 'padding': '10px'}
                ),
                width={'size':8, 'offset': 2}  # Center the card on the page
            )
        ),




        
        # Add the new upload component for the second file
            dbc.Row(
                dbc.Col(
                    dcc.Upload(
                        id='upload-data-compare',
                        children=html.Div([
                            'Drag and Drop or ',
                            html.A('Select Second Excel File for Comparison')
                        ]),
                        style={
                            'width': '100%',
                            'height': '60px',
                            'lineHeight': '60px',
                            'borderWidth': '1px',
                            'borderStyle': 'dashed',
                            'borderRadius': '5px',
                            'textAlign': 'center'
                        },
                        multiple=False
                    ),
                    width=12
                )
            ),
        # Container for displaying comparison results


        dbc.Row([
            dbc.Col(
                dcc.Loading(
                    id="loading-analysis",
                    type="circle",  # Loading style
                    children=[
                        # Wrap the analysis text inside a dbc.Card and dbc.CardBody for a polished look
                        dbc.Card(
                            dbc.CardBody(
                                [
                                    html.Div(id='analysis-text', className="card-text" )  # Use className "card-text" for styling within the card
                                ]
                            ),
                            className="mb-3 shadow",
                                style={'backgroundColor': '#f5fafc', 'borderRadius': '15px', 'padding': '10px'}
                        )
                    ]
                ),
                width=3,  # Half width of the row (12 columns grid system)
                style={'paddingRight': '10px','paddingTop': '30px'}  # Optional: Adds some spacing between columns
            ),
            # Include the other half of the row here if necessary
       
                dbc.Col(
                    dcc.Loading(
                        id="loading-table",
                        type="circle",  # Loading style
                        children=[
                            html.Div(id='comparison-table', style={'marginTop': '50px'}) , # Container for the table
                                        # Add the download button and link container
                            html.Div([
                                    html.Div(id='download-link-container2'),
                                    dbc.Button('Export Data', id='download-button2', color='primary', className='mt-3'),
                                    dcc.Download(id='download-dataframe-csv2')
                                    
                                ], className='text-center'),
                        ]
                    ),
                    width=9,  # Half width of the row
                    style={'paddingLeft': '10px'}  # Optional: Adds some spacing between columns
                )
            ],
           # no_gutters=True,  # Set to False if you want gutters (spacing) between columns
            className="mb-4"  # Optional: Adds margin bottom
        ),
        

    ]
)
app.layout = html.Div([navbar, main_content])
# Callback to handle the loading state and apply blur effect
@app.callback(
    Output('dropdown-container', 'style'),
    [Input('regulatory-filter', 'options')]
)
def update_loading_style(loading_state):
    print('loading ...')
    if not loading_state :
        return {'filter': 'blur(2px)', 'transition': 'filter 0.3s ease'}  # Apply blur when loading
    return {'filter': 'blur(0px)', 'transition': 'filter 0.3s ease'}  # No blur when not loading


# Callback to handle the file upload and display the data
@app.callback(

    [Output('msg_table', 'children'),
     Output('regulatory-filter', 'options'),
     Output('ApplicationRate-filter', 'options')],
    [Input('upload-data', 'contents')],
    [State('upload-data', 'filename')]
)

def data_information(contents,filename):
    global cgap_df 
    if contents is not None:
        
        print('----file not empty------------')
        df=read_file(contents,filename)
        df.columns = df.columns.to_series().apply(
                lambda x: x.replace("rate", "").replace("Max single", "Application rate") 
                if x.startswith("Max single") and x.endswith("(g/ha)") else x
            )
        rate_columns = [
                            {'label': col.replace('\n', ''), 'value': col.replace('\n', '')}
                            for col in df.columns 
                            if( col.startswith("Application rate") or (col.startswith("Max single ") ) and col.endswith("(g/ha)"))
                        ]
        
   
        print(rate_columns)
        zone_list= ['Regulatory Zone','Residue region']
        region_columns= [{'label': col, 'value': col} for col in zone_list]


       # print(region_columns)
        #print(zone_list)

        cgap_df=data_harmonization(df)

        # Return the msg-table with the HTML message and the dropdown options
        return (
            html.Div(['Excel file imported. Select the cGap columns you wish to use for the calculations'],style={'color': 'grey'} ),
        
            region_columns,
            rate_columns

        )

    else:
        return(
            html.Div(['Please upload an Excel file and wait for 5 seconds']),  # Return a tuple for the msg_table
            [],  # Return an empty list for the  options
            []
        )



@app.callback(
    Output('crop-filter', 'options'),
     Output('product-filter', 'options'),
     Output('region-filter', 'options'),
    [Input('regulatory-filter', 'value')],
     [State('upload-data', 'contents'),
     State('upload-data', 'filename')]
)
def update_filter_dropdown(region_columns,contents, filename):
    
    global cgap_df

    if contents is not None and region_columns is not None:
        print('---- updating filter dropdown ---- ')
        print('region_columns is :  ',region_columns)

        print(cgap_df.columns)
        print(cgap_df['Product'].unique())
    

        # Define the options for the dropdowns with the "All" option
        crop_options = [{'label': 'All', 'value': 'All'}] + [{'label': crop, 'value': crop} for crop in cgap_df['Crop'].unique()]
        product_options = [{'label': 'All', 'value': 'All'}] + [{'label': product, 'value': product} for product in cgap_df['Product'].dropna().unique()]
        region_options = [{'label': 'All', 'value': 'All'}] + [{'label': region, 'value': region} for region in cgap_df[region_columns].unique()]
        print(region_options)
        return (
                crop_options,
                product_options,
                region_options,
        )
    else:
        return(
            [],
            [],  # Return an empty list for the  options
            []
        )


@app.callback(
    Output('filtered-table', 'children'),
    [Input('regulatory-filter', 'value'),
     Input('ApplicationRate-filter', 'value'),
     Input('product-filter', 'value'),
     Input('crop-filter', 'value'),
     Input('region-filter', 'value')],
    [State('upload-data', 'contents'),
     State('upload-data', 'filename')]
)


def display_data(region_columns,rate_columns,product_options, crop_options, region_options, contents, filename):
    global critical_values
    global cgap_df 
    

    if contents is not None:
        print('----- function display data triggered-------')       
           
        if rate_columns is not None and region_columns is not None:

            # Filter out rows where PHI is "as per growth stage"
            # Group by and aggregate
            critical_values = (
                cgap_df.groupby([region_columns, 'Product', 'Crop', 'Max # of applns'])
                .agg({
                    rate_columns: 'max',
                    'BBCH latest': 'max',
                    'PHI': lambda x: x[x != 'as per growth stage'].min() if not x[x != 'as per growth stage'].empty else 'as per growth stage',
                    'Interval (Days)': 'min'
                })
                .reset_index()
            )


            print('----- columns selected -------')
            # Add critical flags
            # Calculate critical flags
            critical_values = calculate_critical_flag(critical_values,rate_columns,region_columns)

            # Define the options for the dropdowns with the "All" option
            print(' got here final function')
            # Process the uploaded file and extract options for product and application filters
            print('++++++region_columns++++',region_columns)
            print('++++++rate_columns++++',rate_columns)
            # Update the options for the dropdowns
            print('+++++product_options+++++',product_options)
            print('+++++crop_options+++++',crop_options)
            print('+++++region_options+++++',region_options)
            filtered_values = critical_values
            # Apply filtering based on product and application filters
          
            if product_options is not None:
                if product_options == ['All'] or product_options == []:
                    # Include all available crop options in the filtering process
                    filtered_values = filtered_values[filtered_values['Product'].isin(filtered_values['Product'].unique())]
                else:
                    filtered_values = filtered_values[filtered_values['Product'].isin(product_options)]
            if region_options is not None:
                if region_options == ['All'] or  region_options == [] :
                    # Include all available crop options in the filtering process
                    filtered_values = filtered_values[filtered_values[region_columns].isin(filtered_values[region_columns].unique())]
                else:

                    filtered_values = filtered_values[filtered_values[region_columns].isin(region_options)]
            if crop_options is not None:
                if crop_options == ['All'] or crop_options == []  :
                    # Include all available crop options in the filtering process
                    filtered_values = filtered_values[filtered_values['Crop'].isin(filtered_values['Crop'].unique())]
                else:
                    filtered_values = filtered_values[filtered_values['Crop'].isin(crop_options)]
            
            print(filtered_values.shape)
            print(filtered_values.columns)
          

            print('--------------------------------------------')
          

            

            # Display the filtered dataframe using dash_table.DataTable
            return dash_table.DataTable(
                columns=[{'name': col, 'id': col} for col in filtered_values.columns],
                data=filtered_values.to_dict('records'),
                style_table={
                    'overflowX': 'scroll',
                    'overflowY': 'scroll',
                    'maxHeight': '100vh',
                    'height': '80%',
                    'minWidth': '100%'
                },
                style_cell={
                    'minWidth': '80px', 'maxWidth': '180px', #'whiteSpace': 'normal',
                    'textAlign': 'left',
                        'padding': '5px',
                        'fontFamily': 'Arial, sans-serif',
                        'fontSize': '14px',
                },
                style_header={
                    'backgroundColor':'#f8f9fa',
                    'fontWeight': 'bold',
                        'textAlign': 'center',
                    'whiteSpace': 'normal',
                    'height': 'auto',
                    'maxWidth': '200px',
                },
                style_data_conditional= [{
                                        'if': {
                                            'column_id': col
                                        },
                                        'backgroundColor': 'lightgrey',
                                    } for col in ['is_critical','is_most_critical']],
                                    page_size=20,  # Show 20 rows per page
                    filter_action='native',  # Allow filtering
                    sort_action='native',  # Allow sorting
                page_action='none',
                fixed_rows={'headers': True}
            )
        
        return(
            html.Div(
        ['Select the right columns for the appropriate calculations to be performed.'],
        style={'color': 'red'}  # Set the color to red),  # Return a tuple for the msg_table
            
        ))
        
    

# Callback to generate the download link
@app.callback(
    Output('download-link-container', 'children'),
    Input('download-button', 'n_clicks'),
    prevent_initial_call=True
)

def generate_download_link(n_clicks):
    global critical_values  # Access the global variable
    if n_clicks is not None and n_clicks > 0:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            # Write the DataFrame to Excel
            critical_values.to_excel(writer, index=False, sheet_name='Sheet1')

            # Access the workbook and the sheet
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Enable AutoFilter on the header row
            worksheet.auto_filter.ref = worksheet.dimensions

            # Define the fill styles
            highlight_fill_critical = openpyxl.styles.PatternFill(start_color='ffffe6', end_color='ffffe6', fill_type='solid')  # Light Yellow
            highlight_fill_most_critical = openpyxl.styles.PatternFill(start_color='ffd7c7', end_color='ffd7c7', fill_type='solid')  # Light Orange
            highlight_non_critical = openpyxl.styles.PatternFill(start_color='e1e1e1', end_color='e1e1e1', fill_type='solid')
            # Define the bold border style
            bold_border = Border(left=Side(style='thick'))

            # Apply the highlight to the critical rows
            for row in range(2, len(critical_values) + 2):  # Starting from 2 to account for the header row
                if critical_values['is_critical'].iloc[row - 2]:  # Adjusting index for DataFrame
                    for col in range(1, len(critical_values.columns) + 1):  # Loop through columns
                        worksheet.cell(row=row, column=col).fill = highlight_fill_critical
                else :
                    for col in range(1, len(critical_values.columns) + 1):  # Loop through columns
                        worksheet.cell(row=row, column=col).fill = highlight_non_critical
                try:

                    if critical_values['is_most_critical'].iloc[row - 2]:  # Adjusting index for DataFrame
                        for col in range(len(critical_values.columns)-2, len(critical_values.columns) + 1):  # Loop through columns
                            worksheet.cell(row=row, column=col).fill = highlight_fill_most_critical
                except:
                    print('no most critical values')
                # Add a bold line at the specified column
                column_index = len(critical_values.columns) - 2
                for row in range(2, len(critical_values) + 2):  # Starting from 2 to account for the header row
                    worksheet.cell(row=row, column=column_index).border = bold_border

        output.seek(0)

        # Encode the Excel file to a base64 string
        excel_string = base64.b64encode(output.getvalue()).decode()
        excel_href = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_string}"

        return html.A('Download Excel Data', href=excel_href, download="cGap_data.xlsx", target='_blank', style={'font-weight': 'bold', 'color': 'red'})

    return no_update




# Callback to generate the download link
@app.callback(
    Output('download-link-container2', 'children'),
    Input('download-button2', 'n_clicks'),
    prevent_initial_call=True
)

def generate_download_link2(n_clicks):
    global sorted_df  # Access the global variable
    if n_clicks is not None and n_clicks > 0:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:

            # Write the DataFrame to Excel
            sorted_df.to_excel(writer, index=False, sheet_name='sorted_df')

            # Access the workbook and the sheet
            workbook = writer.book
            worksheet = writer.sheets['sorted_df']

            # Enable AutoFilter on the header row
            worksheet.auto_filter.ref = worksheet.dimensions

            # Define the fill styles
            highlight_fill_critical = openpyxl.styles.PatternFill(start_color='ffffe6', end_color='ffffe6', fill_type='solid')  # Light Yellow
            highlight_fill_most_critical = openpyxl.styles.PatternFill(start_color='ffd7c7', end_color='ffd7c7', fill_type='solid')  # Light Orange
            highlight_non_critical = openpyxl.styles.PatternFill(start_color='e1e1e1', end_color='e1e1e1', fill_type='solid')
            # Define the bold border style
            bold_border = Border(left=Side(style='thick'))

            # Apply the highlight to the critical rows
            for row in range(2, len(sorted_df) + 2):  # Starting from 2 to account for the header row
                if sorted_df['is_critical'].iloc[row - 2]:  # Adjusting index for DataFrame
                    for col in range(1, len(sorted_df.columns) + 1):  # Loop through columns
                        worksheet.cell(row=row, column=col).fill = highlight_fill_critical
                else :
                    for col in range(1, len(sorted_df.columns) + 1):  # Loop through columns
                        worksheet.cell(row=row, column=col).fill = highlight_non_critical
                try:

                    if sorted_df['is_most_critical'].iloc[row - 2]:  # Adjusting index for DataFrame
                        for col in range(len(sorted_df.columns)-2, len(sorted_df.columns) + 1):  # Loop through columns
                            worksheet.cell(row=row, column=col).fill = highlight_fill_most_critical
                except:
                    print('no most critical values')
                # Add a bold line at the specified column
                column_index = len(sorted_df.columns) - 2
                for row in range(2, len(sorted_df) + 2):  # Starting from 2 to account for the header row
                    worksheet.cell(row=row, column=column_index).border = bold_border

        output.seek(0)

        # Encode the Excel file to a base64 string
        excel_string = base64.b64encode(output.getvalue()).decode()
        excel_href = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_string}"

        return html.A('Download merged Excel Data', href=excel_href, download="cGap_data_merged.xlsx", target='_blank', style={'font-weight': 'bold', 'color': 'red'})

    return no_update







@app.callback(
    [Output('analysis-text', 'children'),
     Output('comparison-table', 'children') ],  # For the textual analysis]
    [State('regulatory-filter', 'value'),
     State('ApplicationRate-filter', 'value')],
    [Input('upload-data-compare', 'contents'),
     Input('upload-data-compare', 'filename')]
)


def display_data2(region_columns,rate_columns, contents2, filename2):
    global critical_values 
    global sorted_df  # Access the global variable
    

    if contents2 is not None:
        print('----- new request-------')   
        print('----file not empty------------')
        df2=read_file(contents2,filename2)
        df2.columns = df2.columns.to_series().apply(
                lambda x: x.replace("rate", "").replace("Max single", "Application rate") 
                if x.startswith("Max single") and x.endswith("(g/ha)") else x)
        cgap_df2=data_harmonization(df2)
        print('#######')
           
        if rate_columns is not None and region_columns is not None:

            # Filter out rows where PHI is "as per growth stage"
            # Group by and aggregate
            critical_values2 = (
                cgap_df2.groupby([region_columns, 'Product', 'Crop', 'Max # of applns'])
                .agg({
                    rate_columns: 'max',
                    'BBCH latest': 'max',
                    'PHI': lambda x: x[x != 'as per growth stage'].min() if not x[x != 'as per growth stage'].empty else 'as per growth stage',
                    'Interval (Days)': 'min'
                })
                .reset_index()
            )


            # Add critical flags
            # Calculate critical flags
            critical_values2 = calculate_critical_flag(critical_values2,rate_columns,region_columns)
            # Merging and finding differences
            merged_df = pd.merge(critical_values, critical_values2, how='outer', indicator=True)

            # Add a new column to indicate the origin of the row
            merged_df['Origin'] = merged_df['_merge'].replace({
                'left_only': 'Initial Gap Table',
                'right_only': 'New Gap Table',
                'both': 'both'
            }) 
            changes_only_df = merged_df[merged_df['_merge'] != 'both'].drop(columns=['_merge'])

            # Sort the DataFrame by specified columns
            sorted_df = merged_df.sort_values(by=[rate_columns, 'Product', 'Crop', 'Max # of applns'])

            # Reset index if you want a clean index after sorting
            sorted_df = sorted_df.reset_index(drop=True).drop(columns=['_merge'])





            shape_analysis_text = (
            f"Initial Gap Table has {critical_values.shape[0]} rows.                   \n "
            f"The new Gap Table has {critical_values2.shape[0]}.                       \n"
            f"After merging the 2 Gap Table , the comparison reveals a difference in {changes_only_df.shape[0] } rows                 \n"
            f"where {sorted_df['Origin'].value_counts().get('Initial Gap Table', 0)} rows from the original Table where not kept in the new file.              \n"
             f" and  {sorted_df['Origin'].value_counts().get('New Gap Table', 0)}  rows where added to the new Table.              \n"
            "This analysis helps identify how the new data compares to the initial dataset in terms of its dimensions."
             )             
         

            # Combine the shape analysis text with the summary text
            final_analysis_text =   "\n\nSummary of Changes :\n" + shape_analysis_text 


            if changes_only_df.empty:
                return html.Div("No differences found."), final_analysis_text
            
            # Constructing the DataTable for differences
            data_table = dash_table.DataTable(
                columns=[{'name': col, 'id': col} for col in sorted_df.columns],
                data=sorted_df.to_dict('records'),
                style_table={'overflowX': 'scroll', 'maxHeight': '500px'},
                style_cell={'minWidth': '80px', 'maxWidth': '180px', 'fontSize': '10px'},
                style_header={'backgroundColor': '#f9f9ff', 'textAlign': 'center', 'fontSize': '12px'},
                page_action='none',
                fixed_rows={'headers': True}
            )
            
            return  final_analysis_text,data_table
        else:
            # Handle case where no file is uploaded or contents2 is None
            return html.Div(['Select the right columns for the appropriate calculations to be performed.'],
        style={'color': 'red'}), "Awaiting 2nd file upload..."
   
    return html.Div("Awaiting the upload of your second file..."), ""
                    



''' 


            # Merge both DataFrames with an indicator flag
            merged_df = pd.merge(critical_values, critical_values2, how='outer', indicator=True)

            # Filter the merged DataFrame to include only the rows that are unique to either DataFrame
            changes_only_df = merged_df[merged_df['_merge'] != 'both']

            # Drop the '_merge' column as it's no longer needed
            changes_only_df = changes_only_df.drop(columns=['_merge'])

            print(changes_only_df.shape)
                    

            

            # Display the filtered dataframe using dash_table.DataTable
            return dash_table.DataTable(
                columns=[{'name': col, 'id': col} for col in changes_only_df.columns],
                data=changes_only_df.to_dict('records'),
                style_table={
                    'overflowX': 'scroll',
                    'overflowY': 'scroll',
                    'maxHeight': '100vh',
                    'height': '80%',
                    'minWidth': '80%'
                },
                style_cell={
                    'minWidth': '80px', 'maxWidth': '180px', #'whiteSpace': 'normal',
                    'fontSize': '10px'
                },
                style_header={
                    'backgroundColor':'#f9f9ff',
                    'whiteSpace': 'normal',
                    'height': 'auto',
                    'textAlign': 'center',
                    'maxWidth': '200px',
                    'fontSize': '12px'
                },
                style_data_conditional= [{
                                        'if': {
                                            'column_id': col
                                        },
                                        'backgroundColor': 'lightgrey',
                                    } for col in ['is_critical','is_most_critical']],
                page_action='none',
                fixed_rows={'headers': True}
            )
        
        return(
            html.Div(
        ['Select the appropirate filter first'],
        style={'color': 'red'}  # Set the color to red),  # Return a tuple for the msg_table
            
        ))
          '''
    

'''
@app.callback(
    Output('comparison-table', 'children'),
    [Input('regulatory-filter', 'value'),
    Input('upload-data-compare', 'contents'),
    Input('region-filter', 'value')],
    State('upload-data-compare', 'filename'),
    State('upload-data', 'contents'),
    State('upload-data', 'filename')
)



def compare_files(contents_new, filename_new,region_columns,rate_columns):
    global cgap_df 
    global compare_cgap
    if contents_new is not None:
        print('----file not empty------------')
         # Process the uploaded file and extract data
        content_type, content_string = contents_new.split(',')
        decoded = base64.b64decode(content_string)
        # Read the Excel file into a pandas DataFrame, selecting the specified sheet and skipping rows
    
        try:
            # Try to read the 'MasterGAP' sheet
            df = pd.read_excel(BytesIO(decoded), sheet_name='MasterGAP', skiprows=6)
        except ValueError:
            # If the sheet doesn't exist, try to read the 'DSA GAP overview' sheet
            df = pd.read_excel(BytesIO(decoded), sheet_name='DSA GAP overview', skiprows=2)
            
        new_cgap_df=data_preporcessing (df)


        critical_values1 = (
                cgap_df.groupby([region_columns, 'Product', 'Crop', 'Max # of applns'])
                .agg({
                    rate_columns: 'max',
                    'BBCH latest': 'max',
                    'PHI': lambda x: x[x != 'as per growth stage'].min() if not x[x != 'as per growth stage'].empty else 'as per growth stage',
                    'Interval (Days)': 'min'
                })
                .reset_index()
            )
        

        critical_values2 = (
                new_cgap_df.groupby([region_columns, 'Product', 'Crop', 'Max # of applns'])
                .agg({
                    rate_columns: 'max',
                    'BBCH latest': 'max',
                    'PHI': lambda x: x[x != 'as per growth stage'].min() if not x[x != 'as per growth stage'].empty else 'as per growth stage',
                    'Interval (Days)': 'min'
                })
                .reset_index()
            )
        print(new_cgap_df.columns)

        # Perform an outer merge to compare the two DataFrames
        merged_df = pd.merge(critical_values1, critical_values2, how='outer', indicator=True)

        # Filter for added lines (new entries)
        #added_lines = merged_df[merged_df['_merge'] == 'right_only']

        # Display the comparison results
        return dash_table.DataTable(
            columns=[{'name': col, 'id': col} for col in merged_df.columns],
            data=merged_df.to_dict('records'),
            style_table={
                'overflowX': 'scroll',
                'overflowY': 'scroll',
                'maxHeight': '100vh',
                'height': '80%',
                'minWidth': '100%'
            },
            style_cell={
                'minWidth': '80px', 'maxWidth': '180px',
                'fontSize': '10px'
            },
            style_header={
                'backgroundColor': '#f9f9f9',
                'whiteSpace': 'normal',
                'height': 'auto',
                'textAlign': 'center',
                'maxWidth': '200px',
                'fontSize': '12px'
            },
            page_action='none',
            fixed_rows={'headers': True}
        )

    return html.Div(['Please upload both files to compare.'])

'''
if __name__ == '__main__':
    app.run(debug=True, port=8080,dev_tools_hot_reload=False)
